#KAT Mekatronik AS
#Haptic Reporter
import sys
#import xlsxwriter
from datetime import date
from openpyxl import load_workbook, workbook
import texttable as tab

#from termcolor import colored, cprint
#print_red_on_cyan = lambda x: cprint(x, 'red', 'on_cyan')

print('\nKAT MEKATRONIK URUNLERI AS', end='\n')
print('\nHAPTIK RAPORLAYICI V.DEMO.1.1 2017/4', end='\n')
cnt=0 #multipurpose counter
index=0 #index for distance & force list
templist = [] #Temporary list to save lines in given file.
prompt = str() #Prompt for reporting

##Reporting Details:
temp_report = 'template.xlsx'
r_row = 7
r_col = 1
reportcnt = 0
report_flag = 0

##Get Date:
today = date.today()
today = today.timetuple()
yr = str(today[0])
m = str(today[1])
d = str(today[2])
date = d+'.'+m+'.'+yr

#tuple(distance) #invalid assignment
#tuple(force) #invalid assignment

distance_list = [] #list: distance values to be stored
force_list = [] #list: corresponding values to be stored
result_list = [] #list for results

P1HL = 6.90
P1LL = 5.10
S2HL = 1.92
S2LL = 1.28
SNPHL = 65
SNPLL = 35


P1 = float() #Maximum force value.
P2 = float() #Minimum force value.
P3 = float() #Where P1 meets P3.
S1 = float() #Distance where max force is reached.
S2 = float() #Distance where min force is reacher
S3 = float() #Distance where P1 is reached again.
Result_P1 = str() #store result.
Result_S1 = str() #store result.
Result_S3 = str() #store result.
Result_SNP = str() #store result.
out = ''


while True:
##    P1OK = 0
##    S1OK = 0
##    S3OK = 0
##    SNPOK = 0
    try:
        if cnt == 0: #if it is the first prompt
            filename = input('\n(Çıkmak için "q" harfine basınız)\nLutfen Numune Dosyası İsmini Giriniz: ')
            if (filename == 'Q')| (filename == 'q'):
                if report_flag == 1:
                    report.save(reportxl)
                break
        else: #if user fails
            filename = input('\n\nLutfen Tekrar Dosya İsmini Giriniz: ')
            if (filename == 'Q')| (filename == 'q'):
                if report_flag == 1:
                    report.save(reportxl)
                break
        sample = filename
        filename = filename + '.tsv'
        file = open(filename, 'r')
        cnt=0
        #Extract values from the given file:
        wholefile = file.readlines()
        for line in wholefile:
            if cnt>0:
                templist = line.split()
                force = eval(templist[2])
                distance= abs(eval(templist[3]))
                force_list.insert(index,force)
                distance_list.insert(index,distance)
                index+=1
            else:
                pass
            cnt=+1
        #End of extraction

        cnt=0 #reset counter
        index=0 #reset counter
        index = 0
        #define P1&S1:
        for member in force_list:
            if member > P1:
                P1 = member 
                max_pos = index
                cnt=0
            elif member < P1:
                cnt+=1
                if cnt == 20:
                    break
            index+=1
        P1 = force_list[max_pos] #assign max force
        S1 = distance_list[max_pos] #assign corresponding distance

        ##if next point is equal to P1, check the maximum (search for 10 poinst beyond):
        if P1 == force_list[max_pos+1]:
            inc = 1
            stop = 1
            S1_init = S1
            while (inc < 11) & (stop):
                if P1 == force_list[max_pos+inc]:
                    S1_init= S1_init + distance_list[max_pos+inc]
                    inc+=1
                else:
                    stop = 0
            S1 = S1_init/(inc)
        
        #define P2&S2:
        cnt=0
        length = (len(force_list)-1)-max_pos #define remaining positions
        index = max_pos
        P2 = P1
        for i in range(length):
            if force_list[index] < P2:
                P2 = force_list[index]
                min_pos = index
                cnt = 0
            elif force_list[index] > P2:
                cnt +=1
                if cnt==5:
                    break
            index+=1
        S2 = distance_list[min_pos]
        cnt=0

        ##if net point is equal to P2, check the maximum (search for 10 poinst beyond):
        if P2 == force_list[max_pos+1]:
            inc = 1
            stop = 1
            S2_init = S2
            while (inc < 11) & (stop):
                if P2 == force_list[max_pos+inc]:
                    S2_init= S2_init + distance_list[max_pos+inc]
                    inc+=1
                else:
                    stop = 0
            S2 = S2_init/(inc)
        ##
        #define S3:
        P1MAX = P1*1.05
        P1MIN = P1*0.95
        index=0
        for member in force_list:
            pos = distance_list[index]
            if pos > S2:
                if (member < P1MAX) & (member > P1MIN):
                    P3 = member #define P1=P3
                    S3 = distance_list[index] #assign S3 distance
                    break
            index+=1
        ##

        #Evaluate Results:
        S3MAX = S2*1.20
        SNP = ((P1-P2)/P1)*100

        if (P1<P1HL) & (P1>P1LL):
            P1OK = True
            out1 = '-'
        else:
            P1OK = False
            out1 = 'NOK' 

        if (S2<S2HL) & (S2>S2LL):
            S2OK = True
            out2 = '-'
        else:
            S2OK = False
            out2 = 'NOK'

        if (S3<S3MAX):
            S3OK = True
            out3 = '-'
        else:
            S3OK = False
            out3 = 'NOK'

        if (SNP<SNPHL) & (SNP>SNPLL):
            SNPOK = True
            out4 = '-'
        else:
            SNPOK = False
            out4 = 'NOK'

        if P1OK & S3OK & SNPOK & S2OK:
            status = 'Olumlu'
        else:
            status = 'Olumsuz'            
        

        ##reporting results:
        #Defining table for results
        result = tab.Texttable()
        header = ['P1','S1', 'P2','S2','P3','S3','SNAP','Sonuc']
        result.header(header)
        result.set_cols_width([8,8,8,8,8,8,8,8])
        result.set_cols_align(['r','r','r','r','r','r','r','r'])
        result.set_cols_valign(['m','m','m','m','m','m','m','m'])
        result.set_chars(['-','|','+','#'])
        row=[P1,S1,P2,S2,P3,S3,SNP,status]
        result.add_row(row)
        row = [out1,'','',out2,'',out3,out4,'']
        result.add_row(row)
        s = result.draw()
        print (s)
        
        ##reporting results
        templist = [sample, P1, S1, P2, S2, P3, S3, SNP, status]
        yn = True
        while yn == True:
            if report_flag == 0:
                prompt = input('\n\nSonuçlar raporlasın mı (E/H)?')
                if prompt == 'E':
                    yn = False
                elif prompt =='H':
                    yn = False
                else:
                    print('\n\nYalnızca büyük harf E veya H girin')                    
            elif report_flag == 1:
                prompt = input('\n\nSonuçlar aynı dosyaya mı raporlansın (E/H)?')
                if prompt == 'E':
                    yn = False
                elif prompt =='H':
                    yn = False
                else:
                    print('\n\nYalnızca büyük harf E veya H girin')     

        if (prompt == 'E') & (report_flag == 0):
            report_flag = 1
            report_name = input('\n\nLutfen Rapor Adı Girin:')
            reportxl= report_name+'.xlsx'            
            report = load_workbook(temp_report)
            sheet = report.active

            sheet['B2'] = date            

            index = 0
            while(r_col<=9):
                sheet.cell(row=r_row, column=r_col).value = templist[index]
                r_col+=1
                index+=1
            index=0
            r_col=1

##            report.save(reportxl)
        elif (prompt == 'E') & (report_flag ==1):
            r_row += 1
            index = 0
            while(r_col<=9):
                sheet.cell(row=r_row, column=r_col).value = templist[index]
                r_col+=1
                index+=1
            index=0
            r_col=1
        elif (prompt == 'H') & (report_flag==1):
            report.save(reportxl)
            report_name = input('\n\nLutfen Rapor Adı Girin:')
            reportxl= report_name+'.xlsx'            
            report = load_workbook(temp_report)
            sheet = report.active

            sheet['B2'] = date            

            index = 0
            while(r_col<=9):
                sheet.cell(row=r_row, column=r_col).value = templist[index]
                r_col+=1
                index+=1
            index=0
            r_col=1            
        elif (prompt == 'H') & (report_flag==0):
            pass            
        P1=0
        S2=0
        S3=0        
        file.close()
        force_list.clear()
        distance_list.clear()
        templist.clear()
        result.reset()
    except FileNotFoundError:
        print('\nHATA: ',filename,'bulunamadı.', end='\n')
        cnt=cnt+1
    
    

